# -*- coding: utf-8 -*-
"""
Created on Fri Jan 15 12:49:18 2021

@author: Franco
"""
"""
def agregaInicial(listadeletras,listadestrings,sheet):
    i = 0
    while i < len(listadeletras):
        sheet[listadeletras[i]+'1'] = listadestrings[i]
        i+=1
"""    
#Funcion que elimina de los excel toda la data cargada anteriormente    
def eliminaInicial(sheet):   
    maxRow = sheet.max_row
    maxColumn = sheet.max_column
    i = 2
    while i<= maxRow:
        h = 1
        while h <= maxColumn:
            sheet[get_column_letter(h)+str(i)] = ""
            h+=1
        i+=1
        
#Agrega al Excel datos del pdf que aparecen una sola vez.Ej: Sociedad a Facturar
def agregaSimple(valor,sheet,letra,libre):
    sheet[letra+str(libre)] = valor
#Agrega al Excel datos del pdf que pueden aparecer varias veces.Ej: Fecha de entrega
def agregaComplejo(lista,sheet,letra,libre):
     
     i = libre
     j= 0
     while j<len(lista):
         sheet[letra+str(i)]= lista[j]
         i+=1
         j+=1
#Busca informacion que esta en una linea.\n es un caracter que crea una nueva linea        
def buscaLinea(text,i):
    while not text[i]=="\n":
        i+=1
    i+=1
    h=i
    while not text[h]=="\n":
        h+=1
    
    return (text[i:h])

#Chequea si un string es una fecha         
def isDate(aString):
    if len(aString)!= 10:
        return False
    if aString[0:2].isdigit() and (aString[2] =="/" or aString[2]==".") and aString[3:5].isdigit() and (aString[5] =="/" or aString[5]==".") and aString[6:10].isdigit:  
        return True
    return False
#A partir de un text y un indice busca el proximo substring que es una fecha
def buscaDate(text,i):
    while not isDate(text[i:i+10]):
        i+=1
    return text[i:i+10]
#A partir de un text y un indice busca el proximo numero
def buscaSiguienteNumero(text,h):
    while not text[h].isdigit():
        h+=1
    j=h
    while text[h].isdigit() or text[h]=="," or text[h]=="." or text[h]=="-":
        h+=1
    tupla = (text[j:h],h)
    return tupla
#A partir de un text y un indice busca el anterior numero
def buscaAnteriorNumero(text,h):
    while not text[h].isdigit():
        h-=1
    j=h
    while text[h].isdigit() or text[h]=="," or text[h]=="." or text[h]=="-":
        h-=1
    tupla = (text[h:j+1],j)
    return tupla
#Funcion que extrae informacion del pdf de PAE
def conviertepae(text,sheet,primerolibre):
    posicion = 0
    CUIT=""
    numeroPedido = ""
    fecha = ""
    direccionEntrega=""
    posiciones= []
    numeroProducto = []
    fechaEntrega = []
    descripcion=[]
    cantidad=[]
    precioUnitario=[]
    codigoInspeccion=[]
    precioNeto=[]
    razonSocial = buscaLinea(text,0)
    i = 0
    while i < len(text):
        if i+5 < len(text) and text[i:i+5]=="CUIT:":
            h=i+5
            tupla = buscaSiguienteNumero(text,h)
            CUIT = tupla[0].strip()
        if i+14 < len(text) and text[i:i+14] == "Número pedido:":
            i = i+14
            tupla = buscaSiguienteNumero(text,i)
            numeroPedido = tupla[0].strip()
        if i+16 < len(text) and text[i:i+16]=="Nro.deDocumento:":
            i =i+16
            tupla = buscaSiguienteNumero(text,i)
            numeroPedido = tupla[0].strip()
        if i+6 < len(text) and text[i:i+6]=="Fecha:":
            i = i+6
            fecha = buscaDate(text,i).strip()
        if i+18 < len(text) and text[i:i+18] ==  "DireccióndeEntrega":
            i = i+18
            direccionEntrega = (buscaLinea(text,i))   
        if i+20 < len(text) and text[i:i+20] ==  "Dirección de entrega":
            i = i+20
            direccionEntrega = (buscaLinea(text,i))
        if (i+12 < len(text) and text[i:i+12] == "Precio bruto") or (i+11 < len(text) and text[i:i+11] == "Preciobruto"):
            """
            Se que por cada producto en un momento va a decir 'precio bruto',por
            ende al llegar a ese substring, busco para atras un numero de 12 cifras
            el cual va a ser el numero de producto.
            """
            h = i
            while not text[h-12:h].isdigit():
                h-=1
            posicion+=1
            posiciones.append(str(posicion))
            numeroProducto.append(text[h-12:h])
            j = h
            while not isDate(text[h:h+10]):
                h+=1
            fechaEntrega.append(text[h:h+10])
            juan = text[j:h].strip()
            
            descripcion.append(juan)
            h=h+10
            tupla = buscaSiguienteNumero(text,h)
            cantidad.append(tupla[0])
            h = tupla[1]
            tupla = buscaSiguienteNumero(text,h)
            precioUnitario.append(tupla[0])
            h = tupla[1]
            while not text[h]=="/":
                h+=1
            j=h+1
            while not text[h].isalpha():
                h+=1
            codigoInspeccion.append(text[j:h+1])
            tupla = buscaSiguienteNumero(text,h)
            precioNeto.append(tupla[0])
        i+=1
        variables = [numeroPedido,fecha,razonSocial,CUIT,direccionEntrega,posiciones,cantidad,numeroProducto,descripcion,precioUnitario,fechaEntrega,precioNeto,codigoInspeccion]
    i = 0
    while i < len(variables):
        if i < 5:
            agregaSimple(variables[i],sheet,letras[i],primerolibre)
        else:
            agregaComplejo(variables[i],sheet,letras[i],primerolibre)
        i+=1   
    return len(posiciones)+1
#Funcion que extrae informacion del pdf de YPF
def convierteypf(textoprimerapag,textorestopags,sheet,primerolibre):
    numPedido = ""
    fecha = ""
    RazonSocial = ""
    CUIT=""
    lugardeEntrega =""
    CuitSociedad = ""
    SociedadaFactu = ""
    pos = []
    cant = []
    unidad = []
    material=[]
    valorxunidad=[]
    fechaDeEntrega = []
    text = textoprimerapag
    i = 0
    while i < len(text):
        if i+10 < len(text) and text[i:i+10]=="Núm.pedido":
            tupla = buscaSiguienteNumero(text,i)
            numPedido =tupla[0].strip()
            tupla = buscaDate(text,i)
            fecha = tupla.strip()
        if i+10 < len(text) and text[i:i+11]=="Núm. pedido":
            tupla = buscaSiguienteNumero(text,i)
            numPedido =tupla[0].strip()
            tupla = buscaDate(text,i)
            fecha = tupla.strip()
        if i+12 < len(text) and text[i:i+12]=="RAZON SOCIAL":
            h = i
            while text[h:h+4]!= "CUIT":
                h+=1
            RazonSocial =text[i+15:h]
        if i+4 < len(text) and text[i:i+4]=="CUIT" and CUIT=="":
            h = i
            while text[h:h+9]!= "DIRECCION":
                h+=1
            CUIT =text[i+5:h]
        if i+16 < len(text) and text[i:i+16] == "Lugar de Entrega":
            h = i
            while text[h:h+16] != "Fecha de entrega":
                h+=1
            lugardeEntrega = text[i+17:h]
        if i+17 < len(text) and text[i:i+17]=="Fecha de entrega:":
            if text[i+18].isdigit():
                fechaDeEntrega.append(text[i+18:i+28])
        if i + 32 < len(text) and text[i:i+32] == "Facturar a la orden de YPF S.A -":
            h = i
            while text[h:h+7] != "C.U.I.T":
                h+=1
            SociedadaFactu = text[i+32:h]
            h = h+7
            tupla = buscaSiguienteNumero(text,h)
            CuitSociedad = tupla[0]
            
        i+=1
    copiaLugar = lugardeEntrega
    i = 0
    while i < len(lugardeEntrega):
        if i+25 < len(lugardeEntrega) and lugardeEntrega[i:i+25]=="COMPRA POR CUENTA Y ORDEN":
            h = i+25
            while lugardeEntrega[h:h+4]!="CUIT":
                h+=1
            SociedadaFactu = lugardeEntrega[i+25:h]
            CuitSociedad = lugardeEntrega[h+4:len(lugardeEntrega)]
            copiaLugar = lugardeEntrega[0:i]
            
        i+=1
    lugardeEntrega = copiaLugar        
    text= textorestopags
    h = 0
    while h < len(text):
        """
        Se que por cada pedido la unidad es Metro o und pieza, por ende al llegar
        a un substring asi empieza a buscar todos los datos del pedido.
        """
        if h+5 < len(text) and text[h:h+5] == "Metro" or h+10 < len(text) and text[h:h+10] =="UND./PIEZA":
            if h+5 < len(text) and text[h:h+5] == "Metro":
                unidad.append(text[h:h+5])
            elif  h+10 < len(text) and text[h:h+10] =="UND./PIEZA":
                unidad.append(text[h:h+10])
            tupla = buscaAnteriorNumero(text,h)
            cant.append(tupla[0].strip())
            j=h
            while not(text[j-18:j].isdigit()):
                j-=1
            material.append((text[j-18:j]))
            tupla = buscaAnteriorNumero(text,j-19)
            pos.append(tupla[0])
            tupla = buscaSiguienteNumero(text,h)
            valorxunidad.append(tupla[0].strip())
        if h+3 < len(text) and text[h:h+3] == "Day":
            tupla = buscaSiguienteNumero(text,h)
            fechaDeEntrega.append(tupla[0].strip())    
        h+=1
    #Busca las descripciones de los productos en el excel de materiales   
    descripciones = []        
    for items in material:
        wbDic = openpyxl.load_workbook('Materiales en CM Tenaris (Tubulares).xlsx')
        sheetDic = wbDic['Hoja1']
        loEncontro = False
        i = 2
        while not loEncontro and i < sheetDic.max_row:
            if sheetDic["A"+str(i)].value== int(items):
                descripciones.append(sheetDic["B"+str(i)].value)
                loEncontro = True
            i+=1
        if not loEncontro:
            descripciones.append(items)
        
    variables = [numPedido,fecha,RazonSocial,CUIT,lugardeEntrega,pos,cant,unidad,material,valorxunidad,fechaDeEntrega,SociedadaFactu,CuitSociedad,descripciones]
    #Carga los datos en el excel
    i = 0
    while i < len(variables):
        if i < 5 or (i >10 and i <13):
            agregaSimple(variables[i],sheet,letras[i],primerolibre)
        else:
            agregaComplejo(variables[i],sheet,letras[i],primerolibre)
        i+=1
    return len(pos)+1
    

def returnEntry(arg=None):
    resultLabel.config(text="La ruta no es valida, ingrese otra ruta o presione quit para cerrar el programa")
    result = myEntry.get()
    os.chdir(result.strip())
    #resultLabel.config(text="La ruta es valida,presione quit para comenzar o entre una nueva ruta")
    root.destroy()
    
from tkinter import *
import pdfplumber
import os
import openpyxl
import sys
from openpyxl.utils import get_column_letter, column_index_from_string
#pathIsCorrect = False
path =os.getcwd()
 
root = Tk()

myEntry = Entry(root, width=20)
myEntry.focus()
myEntry.bind("<Return>",returnEntry)
myEntry.pack()
resultLabel = Label(root, text = "")
resultLabel.pack(fill=X) 
enterEntry = Button(root, text= "Enter", command=returnEntry)
enterEntry.pack(fill=X)
root.geometry("+750+400")
buton =Button(root, text="Quit", command=root.destroy)
buton.pack()

root.mainloop()


if(os.getcwd()==path):
    sys.exit(-1)
    
wb = openpyxl.load_workbook('ResumenYPF.xlsx')
wbPae = openpyxl.load_workbook('ResumenPAE.xlsx')               
sheet = wb.active   
sheetPae = wbPae.active
eliminaInicial(sheetPae)
eliminaInicial(sheet)
primerolibre = 2 
primerolibrePae = 2
letras = ["A","B","C","D","E","F","G","H","I","J","K","L","M","N"]
"""
#string = ["Numero de Pedido","Fecha","Razon Social","CUIT","Lugar de entrega","Pos","Cantidad","Unidad","Material","Valor por Unidad","Fecha de entrega","Sociedad a Facturar","CUIT Sociedad a Facturar"]
#stringsPae = ["Numero de Pedido","Fecha","Razon Social","CUIT","Lugar de entrega","Pos","Cantidad","Nro de Producto","Descripcion","Valor por Unidad","Fecha de entrega","Precio Neto","Codigo de Inspeccion"]

#agregaInicial(letras,string,sheet)
#agregaInicial(letras,stringsPae,sheetPae)
"""
#Busca todos los archivos del directorio seleccionado
for file in os.listdir():
    #selecciona solo los pdf
    if file[len(file)-3:len(file)].lower() == 'pdf': 
        with pdfplumber.open(file) as pdf:
            #Determina si es un pdf de YPF O PAE
            esYpf = False
            text = ""
            textorestopags=""
            for page in pdf.pages:
                text+=page.extract_text()
            n = 1
            while n < len(pdf.pages):
                textorestopags+= pdf.pages[n].extract_text()
                n+=1
            i=0
            while i <len(text):
                if i+7 < len(text) and text[i:i+7].lower() =="ypf.com":
                    esYpf = True
                i+=1  
            if esYpf:
                primerolibre += convierteypf(pdf.pages[0].extract_text(),textorestopags,sheet,primerolibre)
            else:
                primerolibrePae+= conviertepae(text,sheetPae,primerolibrePae)
        
           
wb.save('ResumenYPF.xlsx') 
wbPae.save('ResumenPAE.xlsx')
