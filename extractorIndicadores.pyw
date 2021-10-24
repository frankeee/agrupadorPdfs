# -*- coding: utf-8 -*-
"""
Created on Thu Feb 25 16:04:48 2021

@author: Franco
"""
def agregaInicial(sheet):
    if sheet["A1"].value != "Fecha":
        sheet["A1"] = "Fecha"
        sheet["B1"] = "Tipo"
        sheet["C1"] = "Valor"

def agregar(sheet,fecha,string,valor,libre):
    
    Sepuedeagregar = True
    h = 1
    while h<libre and Sepuedeagregar:
        if sheet["B"+str(h)].value == string:
            if sheet["A"+str(h)].value == fecha:
                Sepuedeagregar = False
        h+=1        
    if Sepuedeagregar:
        sheet["A"+str(libre)] = fecha
        sheet["B"+str(libre)] = string
        sheet["C"+str(libre)] = str(valor)
        return libre+1
    return libre        
def buscaUltimoDatoEnExcel(sheet,fila):
    maxColumn = sheet.max_column
    h=1
    ultimo = ""
    while h <= maxColumn:
        if sheet[get_column_letter(h)+str(fila)].value == "Ultimo Dato":
            ultimo = sheet[get_column_letter(h)+str(fila+1)].value
        h+=1
    h = 1
    while h <= maxColumn:
        if sheet[get_column_letter(h)+str(fila)].value == ultimo:
            return (ultimo,sheet[get_column_letter(h)+str(fila+1)].value)
        h+=1
    return "vacio"

def buscaPalabra(text,i):
    h=i
    while text[i].isalpha():
        i+=1
    return text[h:i]

def buscaSiguienteNumero(text,h):
    while not text[h].isdigit():
        h+=1
    j=h
    while text[h].isdigit() or text[h]=="," or text[h]=="." or text[h]=="-":
        h+=1
    tupla = (text[j:h],h)
    return tupla

def buscaIesimoNumero(text,h,i):
    numero = None
    indice = h
    while i >0:
        tupla = buscaSiguienteNumero(text,indice)
        indice = tupla[1]
        numero = tupla[0]
        i-=1
    return numero

def returnEntry(arg=None):
    resultLabel.config(text="La ruta no es valida, ingrese otra ruta o presione quit para cerrar el programa")
    result = myEntry.get()
    os.chdir(result.strip())
    #resultLabel.config(text="La ruta es valida,presione quit para comenzar o entre una nueva ruta")
    root.destroy()

from openpyxl import Workbook
import os.path
from os import path
import requests
import os
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import datetime
import xlrd
import pdfplumber
from tkinter import *
import sys

pat =os.getcwd()
 
root = Tk()

myEntry = Entry(root, width=20)
myEntry.focus()
myEntry.bind("<Return>",returnEntry)
myEntry.pack()
resultLabel = Label(root, text = "")
resultLabel.pack(fill=X) 
enterEntry = Button(root, text= "Enter", command=returnEntry)
enterEntry.pack(fill=X)
root.geometry("+750+400")
buton =Button(root, text="Quit", command=root.destroy)
buton.pack()

root.mainloop()


if(os.getcwd()==pat):
    sys.exit(-1)

date = datetime.datetime.now()

#os.chdir(os.path.join('C:','\\Users','Franco','Documents','datosindicadores'))
wb = None
if path.exists("indicadores.xlsx"):
    wb = openpyxl.load_workbook('indicadores.xlsx')
else:
    wb = Workbook()
    
sheet = wb.active
agregaInicial(sheet)
libre = sheet.max_row
nuevadate= (datetime.datetime.now())

#Entra a la pagina dolar hoy para buscar dolar mayorista y CCL

url = "https://www.dolarhoy.com/"
response = requests.get(url,stream=True)
req = requests.get(url)
soup = BeautifulSoup(req.text, "lxml")
lista = soup.find_all('a')

dolarMayorista = ""
dolarCCL = ""

for item in lista:
    if item.get('href') == "/cotizaciondolarmayorista":
        divs=item.find_all('div')
        for divider in divs:
            if divider.get('class')[0] == "compra":
                #dolar mayorista
                dolarMayorista = (divider.string)
    if item.get('href') == "/cotizaciondolarcontadoconliqui":
        divs=item.find_all('div')
        for divider in divs:
            if divider.get('class')[0] == "compra":
                #dolar ccl
                dolarCCL = (divider.string)

fecha = str(nuevadate.day)+"/"+str(nuevadate.month)+"/"+str(nuevadate.year)
libre = agregar(sheet,fecha,"Dolar Mayorista",dolarMayorista,libre)                
libre = agregar(sheet,fecha,"Dolar CCL",dolarCCL,libre)

 #Busca la url del archivo de indicadores del Indec, prueba primero con la fecha actual y resta de a un dia

loencontro= False
while not loencontro:
    
    mes = nuevadate.month
    if mes<10:
        mes = "0"+str(mes)
    else:
        mes = str(mes)
    
    url = "https://www.indec.gob.ar/ftp/documentos/"+str(nuevadate.year)+mes+str(nuevadate.day)+"_Principales_indicadores_INDEC.xlsx"
    response = requests.get(url,stream=True)
    with open("data.xlsx", "wb") as outfile:
        for chunk in response.iter_content(chunk_size=None):  # Let the server decide.
            outfile.write(chunk)
        
    try:
        wbIndec = openpyxl.load_workbook('data.xlsx')    
        loencontro = True
    except:
        print("no")
    nuevadate = nuevadate - datetime.timedelta(days=1)
    
#Busca datos del Indec en el archivo bajado como data.xlsx

if loencontro:
    wbIndec = openpyxl.load_workbook('data.xlsx')
    sheetIndec = wbIndec.active  
    #inflacion
    tupla = (buscaUltimoDatoEnExcel(sheetIndec,29))
    libre = agregar(sheet,tupla[0],"Inflacion",tupla[1]*100,libre) 
    #POBREZA
    tupla = (buscaUltimoDatoEnExcel(sheetIndec,108))
    libre = agregar(sheet,tupla[0],"Pobreza",tupla[1]*100,libre) 
    #PBI
    tupla = (buscaUltimoDatoEnExcel(sheetIndec,74))
    libre = agregar(sheet,tupla[0] ,"PBI",tupla[1]*100,libre) 
    
    sheetIndec = wbIndec["Sector externo"]
    #Exportaciones
    libre = agregar(sheet,sheetIndec["N11"].value,"Exportaciones",(sheetIndec["N15"].value),libre)
    #Importaciones
    libre = agregar(sheet,sheetIndec["N11"].value,"Importaciones",(sheetIndec["N22"].value),libre)
    #Deuda externa
    libre = agregar(sheet,sheetIndec["N78"].value,"Deuda Externa",(sheetIndec["N81"].value),libre)

#Lo mismo que  el archivo del indec pero para la AFIP

nombreMeses= ["null","enero","febrero","marzo","abril","mayo","junio","julio","agosto","septiembre","octubre","noviembre","diciembre"]
nuevadate= (datetime.datetime.now())
loEncontro = False
i = 0
while not loEncontro and i < 6:
    
    nombredelMes = nombreMeses[nuevadate.month]
    anio = str(nuevadate.year)
    dias = nuevadate.day
        
    url = "http://www.afip.gob.ar/institucional/estudios/archivos/comparativo-"+nombredelMes+"-"+anio+".xls"
    print(url)
    response = requests.get(url,stream=True, timeout=(3.05, 27))
    with open("data2.xls", "wb") as outfile:
        for chunk in response.iter_content(chunk_size=None):  # Let the server decide.
            outfile.write(chunk)
    try:   
        book = xlrd.open_workbook("data2.xls")
        loEncontro = True
    except Exception as e:
        print(e)
    i+=1
    nuevadate = nuevadate - datetime.timedelta(days=dias)

if loEncontro:    
    book = xlrd.open_workbook("data2.xls")
    
    sh = book.sheet_by_index(0)
    numeroFilaTotal = 0
    loEncontro = False
    i = 0
    while i < sh.nrows and not loEncontro:
        if sh.cell_value(rowx=i, colx=1) == "  TOTAL GENERAL" or sh.cell_value(rowx=i, colx=1) == "TOTAL GENERAL":
            #print("si")
            loEncontro = True
            numeroFilaTotal = i
        i+=1
    
        #porcentaje recaudacion interanual
    libre = agregar(sheet,str(nuevadate.day)+"/"+str(nuevadate.month)+"/"+str(nuevadate.year),"Recaudacion Interanual",(sh.cell_value(rowx=numeroFilaTotal, colx=5)),libre)

#Busca en la pagina del BCRA las reservas

url = "http://www.bcra.gov.ar/"
response = requests.get(url,stream=True)
req = requests.get(url)
soup = BeautifulSoup(req.text, "lxml")
lista = soup.find_all('td')
esElquequiero = False
for item in lista:
    if esElquequiero:
        libre = agregar(sheet,str(nuevadate.day)+"/"+str(nuevadate.month)+"/"+str(nuevadate.year),"Reservas",item.string,libre)
        esElquequiero = False
    #print(item.find("b"))
    if item.find("b")!= None:
        if item.find("b").string == "Reservas Internacionales del BCRA":
         esElquequiero = True
         
#Busca la url de la camara del acero

nuevadate= (datetime.datetime.now())

loEncontro = False 
mes = nuevadate.month
if mes<10:
    mes = "0"+str(mes)
else:
    mes = str(mes)
anio = str(nuevadate.year)
i=0
while not loEncontro and i < 6:
    ultimoMesPublicado = nuevadate.month
    url = "http://www.acero.org.ar/wp-content/uploads/"+anio+"/"+mes+"/Produccion-Siderurgica-Argentina-1960-"+anio+"-1-"+str(ultimoMesPublicado)+".pdf"
    print(url)
    response = requests.get(url,stream=True)
    with open("data.pdf", "wb") as outfile:    
        for chunk in response.iter_content(chunk_size=None):  # Let the server decide.
            outfile.write(chunk)
    i+=1
    try:
        with pdfplumber.open('data.pdf') as pdf:
            print("ok")
        loEncontro = True
    except Exception as e:
        print(e)
        nuevadate = nuevadate - datetime.timedelta(days=28)

        
        

if loEncontro:    
    text = ""
    nuevadate= (datetime.datetime.now())
    with pdfplumber.open('data.pdf') as pdf:
        
        for page in pdf.pages:
            text += page.extract_text()
    
    
        
    anio = str(nuevadate.year)
    mes = ""
    AceroCrudo = ""
    laminadosNoPlanos = ""
    i = len(text)-1
    while i>=0:
        if text[i] == "\n":
            if text[i+1:i+5].isdigit():
                anio = text[i+1:i+5]
            else:
                if buscaIesimoNumero(text,i,4)!= "0,0":
                    mes = buscaPalabra(text,i+1)
                    AceroCrudo = buscaIesimoNumero(text,i,4)
                    laminadosNoPlanos = buscaIesimoNumero(text,i,5)
                    i = 0
        i-=1
    i = 1
    while i < len(nombreMeses):
        if nombreMeses[i]== mes:
            mes = i
            break
        i+=1
    libre = agregar(sheet,"01/"+str(mes)+"/"+anio,"Acero Crudo",AceroCrudo,libre)
    libre = agregar(sheet,"01/"+str(mes)+"/"+anio,"Laminados no planos",laminadosNoPlanos,libre)


wb.save('indicadores.xlsx')

    